import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color Constants matching user sheet
DARK_BLUE   = '1F3864'
MED_BLUE    = '2E75B6'
LIGHT_BLUE  = 'D9E2F3'
ACCENT_BLUE = 'BDD7EE'
WHITE       = 'FFFFFF'
LIGHT_GRAY  = 'F2F2F2'
GREEN_BG    = 'E2EFDA'
ORANGE_BG   = 'FCE4D6'
RED_BG      = 'FFDDD8'

def parse_date(date_str):
    if not date_str or str(date_str).strip() in ["-", "N/A", "N/A / N/A"]:
        return None
    # Try different formats
    formats = [
        "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d-%b-%Y",
        "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d"
    ]
    date_str_clean = str(date_str).strip()
    for fmt in formats:
        try:
            return datetime.strptime(date_str_clean, fmt)
        except ValueError:
            continue
    # Try regex matching like 12-04-2022
    match = re.search(r"(\d{2})[-\/](\d{2})[-\/](\d{4})", date_str_clean)
    if match:
        try:
            return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except ValueError:
            pass
    # Try 2 digit year
    match = re.search(r"(\d{2})[-\/](\d{2})[-\/](\d{2})", date_str_clean)
    if match:
        try:
            year = int(match.group(3))
            year = 2000 + year if year < 80 else 1900 + year
            return datetime(year, int(match.group(2)), int(match.group(1)))
        except ValueError:
            pass
    return None

def extract_monthly_income(income_band_str, default=0.0) -> float:
    if not income_band_str or str(income_band_str).strip() in ["-", "N/A", ""]:
        return default
    
    ib_clean = str(income_band_str).lower()
    
    # Try to extract numbers
    numbers = re.findall(r"(\d+[\d\,\.]*)", ib_clean)
    if not numbers:
        return default
        
    # Take the first number found and parse it
    val_str = numbers[0].replace(",", "")
    try:
        val = float(val_str)
    except ValueError:
        return default
        
    # Check if value is in Lakhs / Lacs
    is_lakh = any(x in ib_clean for x in ["lakh", "lac", "lacs"])
    
    if is_lakh:
        annual_income = val * 100000
    elif val < 1000:
        try:
            annual_income = float(val_str) * 100000
        except ValueError:
            annual_income = val
    else:
        annual_income = val
        
    # If the annual income is too small to be annual, treat it as monthly
    if annual_income < 20000:
        return annual_income
        
    return annual_income / 12.0

def safe_num(val) -> float:
    if not val or str(val).strip() in ["-", "N/A", "N/A / N/A"]: 
        return 0.0
    cleaned = re.sub(r"[^\d\.]", "", str(val))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0

def to_float(val):
    if not val or str(val).strip() in ["-", "N/A", "N/A / N/A", ""]: 
        return "-"
    cleaned = re.sub(r"[^\d\.]", "", str(val))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return str(val)

def safe(v):
    """Convert any value safely for Excel cells."""
    if v is None or v == {} or v == []: return '-'
    if isinstance(v, dict):             return '-'
    if isinstance(v, list):             return ', '.join(str(x) for x in v)
    return v

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(c, text, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
    """Dark header cell — bold text with specific background."""
    c.value = text
    c.font      = Font(name='Arial', bold=bold, color=fg, size=size)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def lbl(c, text):
    """Blue label cell (left column)."""
    c.value     = text
    c.font      = Font(name='Arial', bold=True, size=10, color='1F3864')
    c.fill      = PatternFill('solid', start_color=LIGHT_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')

def val(c, v, bold=False, center=False, bg=None, color='000000', number_format=None):
    """Value cell — optional background, text color, and format."""
    c.value     = safe(v)
    c.font      = Font(name='Arial', size=10, bold=bold, color=color)
    if bg:
        c.fill  = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center', wrap_text=True
    )
    if number_format and isinstance(v, (int, float)):
        c.number_format = number_format

def bdr(ws, r1, r2, c1, c2):
    """Apply thin border to a rectangular range."""
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = thin_border()

def generate_excel_report(data: dict, ai_result: dict = None) -> bytes:
    """Generates the structured Excel sheet to match user shared analysis."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Extract data variables
    personal = data.get("personal_info", {})
    metadata = data.get("metadata", {})
    score = data.get("credit_score")
    metrics = data.get("summary_metrics", {})
    accounts = data.get("accounts", [])
    inquiries = data.get("inquiries", [])
    
    # Calculate derived dates and statistics
    report_date = parse_date(metadata.get("report_date")) or datetime.now()
    
    # Accounts stats
    total_outstanding = 0.0
    total_overdue = 0.0
    total_limit = 0.0
    total_disbursed = 0.0
    
    secured_count = 0
    unsecured_count = 0
    written_off_count = 0
    settled_count = 0
    restructured_count = 0
    
    parsed_dates = []
    
    secured_keywords = ["home", "housing", "property", "auto", "vehicle", "car", "gold", "two wheeler", "motorcycle", "secured"]
    
    for acct in accounts:
        acct_type = str(acct.get("Account Type", "")).lower()
        status = str(acct.get("Status", "")).lower()
        
        # Secured vs Unsecured count
        if any(re.search(rf"\b{re.escape(kw)}\b", acct_type) for kw in secured_keywords):
            secured_count += 1
        else:
            unsecured_count += 1
            
        # Status counts
        if "write" in status or "written" in status:
            written_off_count += 1
        if "settled" in status:
            settled_count += 1
        if "restructured" in status:
            restructured_count += 1
            
        # Amount sums
        curr_bal = safe_num(acct.get("Current Balance"))
        overdue_amt = safe_num(acct.get("Overdue Amount"))
        credit_limit = safe_num(acct.get("Credit Limit/Sanctioned"))
        high_credit = safe_num(acct.get("High Credit"))
        
        total_outstanding += curr_bal
        total_overdue += overdue_amt
        total_limit += credit_limit
        total_disbursed += high_credit if high_credit > 0 else credit_limit
        
        # Age
        opened_dt = parse_date(acct.get("Date Opened"))
        if opened_dt:
            parsed_dates.append(opened_dt)
            
    # Calculate history and average age
    avg_age_str = "-"
    history_str = "-"
    new_accounts_6m = 0
    
    if parsed_dates:
        ages = []
        for d in parsed_dates:
            diff_months = (report_date.year - d.year) * 12 + (report_date.month - d.month)
            diff_months = max(0, diff_months)
            ages.append(diff_months)
            if diff_months <= 6:
                new_accounts_6m += 1
                
        avg_months = sum(ages) / len(ages)
        avg_age_str = f"{int(avg_months // 12)} yrs {int(avg_months % 12)} months"
        
        earliest_date = min(parsed_dates)
        history_months = (report_date.year - earliest_date.year) * 12 + (report_date.month - earliest_date.month)
        history_months = max(0, history_months)
        history_str = f"{int(history_months // 12)} yrs {int(history_months % 12)} months"

    # Inquiries count last 6 months
    inquiries_6m = 0
    for inq in inquiries:
        inq_dt = parse_date(inq.get("Inquiry Date"))
        if inq_dt:
            diff_months = (report_date.year - inq_dt.year) * 12 + (report_date.month - inq_dt.month)
            if diff_months <= 6:
                inquiries_6m += 1

    # Update metrics fallback if they are empty
    metrics["total_accounts"] = metrics.get("total_accounts") or len(accounts)
    metrics["active_accounts"] = metrics.get("active_accounts") or sum(1 for a in accounts if "active" in str(a.get("Status")).lower())
    metrics["closed_accounts"] = metrics.get("closed_accounts") or sum(1 for a in accounts if "closed" in str(a.get("Status")).lower())
    metrics["overdue_accounts"] = metrics.get("overdue_accounts") or sum(1 for a in accounts if safe_num(a.get("Overdue Amount")) > 0)
    metrics["total_outstanding"] = metrics.get("total_outstanding") or total_outstanding
    metrics["total_overdue"] = metrics.get("total_overdue") or total_overdue

    # -------------------------------------------------------------------------
    # SHEET 1 — Applicant Overview
    # -------------------------------------------------------------------------
    ws1 = wb.create_sheet('Applicant Overview')
    ws1.sheet_view.showGridLines = False
    
    # Title Banner
    ws1.merge_cells('A1:F1')
    hdr(ws1['A1'], 'CREDIT BUREAU REPORT  —  APPLICANT OVERVIEW', size=13)
    ws1.row_dimensions[1].height = 34
    
    # Report meta info
    ws1.merge_cells('A2:F2')
    ws1.row_dimensions[2].height = 20
    meta_c = ws1['A2']
    meta_c.value = (
        f"Report Reference No: {metadata.get('ref_number', 'N/A')}   |   "
        f"Prepared For: {personal.get('Name', 'N/A')}   |   "
        f"Date of Issue: {metadata.get('report_date', 'N/A')}"
    )
    meta_c.font      = Font(name='Arial', size=9, italic=True, color='595959')
    meta_c.fill      = PatternFill('solid', start_color=ACCENT_BLUE)
    meta_c.alignment = Alignment(horizontal='center', vertical='center')
    
    # Section: Personal Information
    ws1.merge_cells('A4:F4')
    hdr(ws1['A4'], 'PERSONAL INFORMATION', bg=MED_BLUE)
    ws1.row_dimensions[4].height = 22
    
    personal_rows = [
        ('Full Name',         personal.get('Name', 'N/A')),
        ('Father\'s Name',     personal.get('Father', 'N/A')),
        ('Date of Birth',     personal.get('DOB', 'N/A')),
        ('Gender',            personal.get('Gender', 'N/A')),
        ('Phone',             personal.get('Mobile', 'N/A')),
        ('PAN',               personal.get('PAN', 'N/A')),
        ('Aadhaar (masked)',  personal.get('Aadhaar', 'N/A')),
        ('Address',           personal.get('Address', 'N/A')),
    ]
    for idx, (label_txt, value_txt) in enumerate(personal_rows):
        r = 5 + idx
        ws1.row_dimensions[r].height = 18
        ws1.merge_cells(f'A{r}:B{r}')
        lbl(ws1[f'A{r}'], label_txt)
        ws1.merge_cells(f'C{r}:F{r}')
        val(ws1[f'C{r}'], value_txt)
    bdr(ws1, 4, 12, 1, 6)
    
    # Section: Report Inquiry Details
    ws1.merge_cells('A14:F14')
    hdr(ws1['A14'], 'REPORT AND SCORE AUDIT DETAILS', bg=MED_BLUE)
    ws1.row_dimensions[14].height = 22
    
    audit_rows = [
        ('Reference Identification', metadata.get('ref_number', 'N/A')),
        ('Date Prepared',            metadata.get('report_date', 'N/A')),
        ('Total Accounts Audited',   metrics.get('total_accounts', 0)),
        ('Lender Inquiries Count',   len(inquiries)),
    ]
    for idx, (label_txt, value_txt) in enumerate(audit_rows):
        r = 15 + idx
        ws1.row_dimensions[r].height = 18
        ws1.merge_cells(f'A{r}:B{r}')
        lbl(ws1[f'A{r}'], label_txt)
        ws1.merge_cells(f'C{r}:F{r}')
        val(ws1[f'C{r}'], value_txt)
    bdr(ws1, 14, 18, 1, 6)
    
    # Section: Credit Score details
    ws1.merge_cells('A20:F20')
    hdr(ws1['A20'], 'CREDIT BUREAU SCORES', bg=MED_BLUE)
    ws1.row_dimensions[20].height = 22
    
    rating = "No Score / NA"
    score_color = '000000'
    bg_color = None
    if score:
        if score >= 750:
            rating = "EXCELLENT"
            score_color = '27AE60'
            bg_color = GREEN_BG
        elif score >= 650:
            rating = "GOOD"
            score_color = '27AE60'
            bg_color = GREEN_BG
        elif score >= 550:
            rating = "FAIR"
            score_color = 'F39C12'
            bg_color = ORANGE_BG
        else:
            rating = "POOR"
            score_color = 'C00000'
            bg_color = RED_BG

    score_rows = [
        ('CRIF High Mark Score', str(score) if score else "N/A", True, score_color, bg_color),
        ('Score Rating Standing', rating, True, score_color, bg_color),
        ('Score Parameters', "Score ranges from 300 to 900. Higher is better.", False, '595959', None)
    ]
    for idx, (label_txt, value_txt, bold, text_col, bg_col) in enumerate(score_rows):
        r = 21 + idx
        ws1.row_dimensions[r].height = 18
        ws1.merge_cells(f'A{r}:B{r}')
        lbl(ws1[f'A{r}'], label_txt)
        ws1.merge_cells(f'C{r}:F{r}')
        val(ws1[f'C{r}'], value_txt, bold=bold, color=text_col, bg=bg_col)
    bdr(ws1, 20, 23, 1, 6)
    
    # Set exact widths
    for col, w in zip(['A','B','C','D','E','F'], [18, 18, 22, 16, 12, 12]):
        ws1.column_dimensions[col].width = w

    # -------------------------------------------------------------------------
    # SHEET 2 — Accounts Summary
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet('Accounts Summary')
    ws2.sheet_view.showGridLines = False
    
    # Title
    ws2.merge_cells('A1:E1')
    hdr(ws2['A1'], 'ACCOUNTS SUMMARY', size=13)
    ws2.row_dimensions[1].height = 32
    
    # Primary accounts
    ws2.merge_cells('A3:E3')
    hdr(ws2['A3'], 'PRIMARY ACCOUNTS', bg=MED_BLUE)
    ws2.row_dimensions[3].height = 22
    
    primary_rows = [
        ('Total Accounts',       metrics.get('total_accounts', 0)),
        ('Active Accounts',      metrics.get('active_accounts', 0)),
        ('Closed Accounts',      metrics.get('closed_accounts', 0)),
        ('Overdue Accounts',     metrics.get('overdue_accounts', 0)),
        ('Secured Accounts',     secured_count),
        ('Unsecured Accounts',   unsecured_count),
        ('Sanctioned/Limit Amt',  f"Rs. {int(total_limit):,}" if total_limit > 0 else "-"),
        ('Disbursed/High Amt',   f"Rs. {int(total_disbursed):,}" if total_disbursed > 0 else "-"),
        ('Current Balance',      f"Rs. {int(safe_num(metrics.get('total_outstanding', 0))):,}" if safe_num(metrics.get('total_outstanding', 0)) > 0 else "-"),
        ('Total Overdue Balance', f"Rs. {int(safe_num(metrics.get('total_overdue', 0))):,}" if safe_num(metrics.get('total_overdue', 0)) > 0 else "-"),
    ]
    for idx, (label_txt, value_txt) in enumerate(primary_rows):
        r = 4 + idx
        ws2.row_dimensions[r].height = 18
        lbl(ws2[f'A{r}'], label_txt)
        ws2.merge_cells(f'B{r}:E{r}')
        val(ws2[f'B{r}'], value_txt, center=True)
    bdr(ws2, 3, 13, 1, 5)
    
    # Derived attributes
    ws2.merge_cells('A16:E16')
    hdr(ws2['A16'], 'DERIVED ATTRIBUTES', bg=MED_BLUE)
    ws2.row_dimensions[16].height = 22
    
    derived_rows = [
        ('Average Account Age',            avg_age_str),
        ('Length of Credit History',       history_str),
        ('Total Written-off Accounts',     written_off_count),
        ('Total Settled Accounts',         settled_count),
        ('Total Restructured Accounts',    restructured_count),
        ('Inquiries in Last 6 Months',     inquiries_6m),
        ('New Accounts in Last 6 Months',  new_accounts_6m),
    ]
    for idx, (label_txt, value_txt) in enumerate(derived_rows):
        r = 17 + idx
        ws2.row_dimensions[r].height = 18
        lbl(ws2[f'A{r}'], label_txt)
        ws2.merge_cells(f'B{r}:E{r}')
        val(ws2[f'B{r}'], value_txt, center=True)
    bdr(ws2, 16, 23, 1, 5)
    
    for col, w in zip(['A','B','C','D','E'], [32, 16, 16, 16, 16]):
        ws2.column_dimensions[col].width = w

    # Separate Active and Closed Accounts for detailed analysis
    active_loans = [a for a in accounts if "active" in str(a.get("Status", "")).lower()]
    closed_loans = [a for a in accounts if "closed" in str(a.get("Status", "")).lower()]
    
    # -------------------------------------------------------------------------
    # SHEET 3 — Dashboard & Analytical Insights
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet('Dashboard & Insights', 1)
    ws3.sheet_view.showGridLines = False
    
    ws3.merge_cells('A1:G1')
    hdr(ws3['A1'], 'COMPREHENSIVE DASHBOARD & ANALYTICAL INSIGHTS', size=13)
    ws3.row_dimensions[1].height = 34
    
    # Section 1: Active Loans Summary
    ws3.merge_cells('A3:G3')
    hdr(ws3['A3'], 'ACTIVE LOANS ANALYSIS', bg=MED_BLUE)
    ws3.row_dimensions[3].height = 22
    
    active_balance = sum(safe_num(a.get("Current Balance", 0)) for a in active_loans)
    active_limit = sum(safe_num(a.get("Credit Limit/Sanctioned", 0)) for a in active_loans)
    active_overdue = sum(safe_num(a.get("Overdue Amount", 0)) for a in active_loans)
    active_emi = sum(safe_num(a.get("Installment Amount", 0)) for a in active_loans)
    
    active_analysis = [
        ('Total Active Loans', len(active_loans)),
        ('Total Outstanding Balance', f"Rs. {int(active_balance):,}" if active_balance > 0 else "Rs. 0"),
        ('Total Overdue Amount', f"Rs. {int(active_overdue):,}" if active_overdue > 0 else "Rs. 0"),
        ('Total Monthly EMI', f"Rs. {int(active_emi):,}" if active_emi > 0 else "Rs. 0"),
        ('Average Account Age', avg_age_str),
    ]
    
    for idx, (label_txt, value_txt) in enumerate(active_analysis):
        r = 4 + idx
        ws3.row_dimensions[r].height = 18
        ws3.merge_cells(f'A{r}:C{r}')
        lbl(ws3[f'A{r}'], label_txt)
        ws3.merge_cells(f'D{r}:G{r}')
        bg_col = RED_BG if "Overdue" in label_txt and active_overdue > 0 else None
        val(ws3[f'D{r}'], value_txt, center=True, bg=bg_col)
    bdr(ws3, 3, 3 + len(active_analysis), 1, 7)
    
    # Section 2: Closed Loans Summary
    ws3.merge_cells('A13:G13')
    hdr(ws3['A13'], 'CLOSED LOANS ANALYSIS', bg=MED_BLUE)
    ws3.row_dimensions[13].height = 22
    
    closed_balance = sum(safe_num(a.get("Current Balance", 0)) for a in closed_loans)
    closed_limit = sum(safe_num(a.get("Credit Limit/Sanctioned", 0)) for a in closed_loans)
    
    closed_analysis = [
        ('Total Closed Loans', len(closed_loans)),
        ('Total Outstanding Balance', f"Rs. {int(closed_balance):,}" if closed_balance > 0 else "Rs. 0"),
        ('Total Credit Limit/Sanctioned', f"Rs. {int(closed_limit):,}" if closed_limit > 0 else "Rs. 0"),
        ('Closure Percentage', f"{(len(closed_loans)/len(accounts)*100):.2f}%" if len(accounts) > 0 else "0%"),
    ]
    
    for idx, (label_txt, value_txt) in enumerate(closed_analysis):
        r = 14 + idx
        ws3.row_dimensions[r].height = 18
        ws3.merge_cells(f'A{r}:C{r}')
        lbl(ws3[f'A{r}'], label_txt)
        ws3.merge_cells(f'D{r}:G{r}')
        val(ws3[f'D{r}'], value_txt, center=True)
    bdr(ws3, 13, 13 + len(closed_analysis), 1, 7)
    
    # Section 3: Credit Health Indicators
    ws3.merge_cells('A20:G20')
    hdr(ws3['A20'], 'CREDIT HEALTH INDICATORS', bg=MED_BLUE)
    ws3.row_dimensions[20].height = 22
    
    income_band_str = data.get("income_band", "-")
    monthly_income = extract_monthly_income(income_band_str, default=0.0)
    if monthly_income > 0:
        if active_emi > 0:
            util_pct = (active_emi / monthly_income) * 100
            repayment_capacity_str = f"{util_pct:.1f}% of parsed income (₹{int(monthly_income):,}/mo)"
            repayment_status = 'Healthy' if util_pct < 30 else 'Moderate' if util_pct < 50 else 'High'
        else:
            repayment_capacity_str = "N/A (0 EMI)"
            repayment_status = "Healthy"
    else:
        if active_emi > 0:
            repayment_capacity_str = f"₹{int(active_emi):,}/mo (Income not available)"
            repayment_status = "N/A"
        else:
            repayment_capacity_str = "N/A (0 EMI)"
            repayment_status = "Healthy"

    health_indicators = [
        ('Overall Credit Score', f"{score}" if score else "N/A", 'EXCELLENT' if score and score >= 750 else 'GOOD' if score and score >= 650 else 'FAIR' if score and score >= 550 else 'POOR'),
        ('Loan Diversification', f"{len(set(a.get('Account Type', '') for a in active_loans))} types", 'Healthy' if len(set(a.get('Account Type', '') for a in active_loans)) >= 2 else 'Limited'),
        ('Payment Discipline', f"0 Overdue" if active_overdue == 0 else f"Overdue: Rs. {int(active_overdue):,}", 'Excellent' if active_overdue == 0 else 'Needs Attention'),
        ('Debt Repayment Capacity', repayment_capacity_str, repayment_status),
    ]
    
    for idx, (label_txt, value_txt, status_txt) in enumerate(health_indicators):
        r = 21 + idx
        ws3.row_dimensions[r].height = 18
        ws3.merge_cells(f'A{r}:B{r}')
        lbl(ws3[f'A{r}'], label_txt)
        ws3.merge_cells(f'C{r}:E{r}')
        val(ws3[f'C{r}'], value_txt, center=True)
        ws3.merge_cells(f'F{r}:G{r}')
        status_bg = GREEN_BG if status_txt in ['Excellent', 'Healthy'] else ORANGE_BG if status_txt in ['Moderate', 'Good'] else RED_BG if status_txt in ['Needs Attention', 'High'] else None
        val(ws3[f'F{r}'], status_txt, center=True, bg=status_bg, bold=True)
    bdr(ws3, 20, 20 + len(health_indicators), 1, 7)
    
    for col, w in zip(['A','B','C','D','E','F','G'], [20, 20, 20, 20, 14, 16, 16]):
        ws3.column_dimensions[col].width = w

    # -------------------------------------------------------------------------
    # SHEET 4 — Active Loans Detailed
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet('Active Loans', 2)
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A3'
    
    ws4.merge_cells('A1:S1')
    hdr(ws4['A1'], f'ACTIVE LOANS DETAILS  ({len(active_loans)} Active Loans)', size=13, bg='27AE60')
    ws4.row_dimensions[1].height = 32
    
    active_col_headers = [
        '#', 'Lender Name', 'Account Number', 'Account Type', 
        'Date Opened', 'Last Reported', 'Tenure (Months)',
        'Credit Limit/Sanctioned', 'Current Balance', 'Utilization %',
        'Overdue Amount', 'Interest Rate', 'Monthly EMI', 
        'High Credit', 'Asset Classification', 'Security',
        'Ownership', 'Status', 'DPD History'
    ]
    for j, h in enumerate(active_col_headers):
        hdr(ws4.cell(2, j + 1), h, bg=MED_BLUE, size=9)
    ws4.row_dimensions[2].height = 28
    
    for i, acct in enumerate(active_loans):
        rn = i + 3
        ws4.row_dimensions[rn].height = 22
        
        curr_bal_val = safe_num(acct.get("Current Balance", 0))
        limit_val = safe_num(acct.get("Credit Limit/Sanctioned", 0))
        utilization = (curr_bal_val / limit_val * 100) if limit_val > 0 else 0
        
        # Calculate tenure
        opened_dt = parse_date(acct.get("Date Opened", "-"))
        tenure_months = "-"
        if opened_dt:
            tenure_months = (report_date.year - opened_dt.year) * 12 + (report_date.month - opened_dt.month)
            tenure_months = max(0, tenure_months)
        
        overdue_val = safe_num(acct.get("Overdue Amount", 0))
        installment_val = safe_num(acct.get("Installment Amount", 0))
        
        row_vals = [
            i + 1,
            acct.get("Lender", "-"),
            acct.get("Account Number", "-"),
            acct.get("Account Type", "-"),
            acct.get("Date Opened", "-"),
            acct.get("Last Reported", "-"),
            tenure_months,
            limit_val,
            curr_bal_val,
            utilization,
            overdue_val,
            acct.get("Interest Rate", "-"),
            installment_val,
            safe_num(acct.get("High Credit", 0)),
            acct.get("Asset Classification", "-"),
            acct.get("Security", "-"),
            acct.get("Ownership", "-"),
            acct.get("Status", "-"),
            acct.get("DPD History / Payment History", "-")
        ]
        
        for j, v in enumerate(row_vals):
            c = ws4.cell(rn, j + 1)
            c.value = safe(v) if not isinstance(v, (int, float)) else v
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', start_color=GREEN_BG)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
            
            if j in [7, 8, 10, 12, 13] and isinstance(v, (int, float)) and v > 0:
                c.number_format = '₹#,##,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
            elif j == 9 and isinstance(v, (int, float)):
                c.number_format = '0.00%'
                c.alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, w in enumerate([3, 18, 14, 16, 13, 13, 13, 14, 14, 12, 13, 12, 12, 12, 14, 12, 12, 12, 15]):
        col_letter = get_column_letter(col_idx + 1)
        ws4.column_dimensions[col_letter].width = w

    # -------------------------------------------------------------------------
    # SHEET 5 — Closed Loans Detailed
    # -------------------------------------------------------------------------
    ws5 = wb.create_sheet('Closed Loans', 3)
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = 'A3'
    
    ws5.merge_cells('A1:T1')
    hdr(ws5['A1'], f'CLOSED LOANS DETAILS  ({len(closed_loans)} Closed Loans)', size=13, bg='1F4E78')
    ws5.row_dimensions[1].height = 32
    
    closed_col_headers = [
        '#', 'Lender Name', 'Account Number', 'Account Type', 
        'Date Opened', 'Last Reported', 'Tenure (Months)', 'Closure Status',
        'Original Limit', 'Highest Balance', 'Final Balance', 
        'Total Interest Rate', 'Total EMI Paid', 'Written-off Amount',
        'Settlement Status', 'Restructuring Status', 'Asset Classification',
        'Security', 'Ownership', 'Payment History'
    ]
    for j, h in enumerate(closed_col_headers):
        hdr(ws5.cell(2, j + 1), h, bg='7F7F7F' if j % 2 == 0 else '999999', size=9, fg=WHITE)
    ws5.row_dimensions[2].height = 28
    
    for i, acct in enumerate(closed_loans):
        rn = i + 3
        ws5.row_dimensions[rn].height = 22
        
        # Calculate tenure
        opened_dt = parse_date(acct.get("Date Opened", "-"))
        last_rep_dt = parse_date(acct.get("Last Reported", "-"))
        tenure_months = "-"
        if opened_dt and last_rep_dt:
            tenure_months = (last_rep_dt.year - opened_dt.year) * 12 + (last_rep_dt.month - opened_dt.month)
            tenure_months = max(0, tenure_months)
        
        status = str(acct.get("Status", "")).lower()
        closure_status = "Normal Closure" if "closed" in status else ("Settlement" if "settled" in status else "Write-off" if "written" in status or "write" in status else "Unknown")
        
        installment_val = safe_num(acct.get("Installment Amount", 0))
        total_emi_paid = "-"
        if installment_val > 0 and isinstance(tenure_months, int):
            total_emi_paid = installment_val * tenure_months
            
        row_vals = [
            i + 1,
            acct.get("Lender", "-"),
            acct.get("Account Number", "-"),
            acct.get("Account Type", "-"),
            acct.get("Date Opened", "-"),
            acct.get("Last Reported", "-"),
            tenure_months,
            closure_status,
            safe_num(acct.get("Credit Limit/Sanctioned", 0)),
            safe_num(acct.get("High Credit", 0)),
            safe_num(acct.get("Current Balance", 0)),
            acct.get("Interest Rate", "-"),
            total_emi_paid,
            safe_num(acct.get("Write-off Amount", 0)),
            acct.get("WO/Settled Status", "-"),
            "No" if "restructured" not in status else "Yes",
            acct.get("Asset Classification", "-"),
            acct.get("Security", "-"),
            acct.get("Ownership", "-"),
            acct.get("DPD History / Payment History", "-")
        ]
        
        for j, v in enumerate(row_vals):
            c = ws5.cell(rn, j + 1)
            c.value = safe(v) if not isinstance(v, (int, float)) else v
            c.font = Font(name='Arial', size=9)
            
            # Determine background color based on closure status
            if closure_status == "Write-off":
                cell_bg = RED_BG
            elif closure_status == "Settlement":
                cell_bg = ORANGE_BG
            else:
                cell_bg = LIGHT_GRAY
            
            c.fill = PatternFill('solid', start_color=cell_bg)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
            
            if j in [8, 9, 10, 13] and isinstance(v, (int, float)) and v > 0:
                c.number_format = '₹#,##,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
    
    for col_idx, w in enumerate([3, 16, 14, 16, 13, 13, 13, 14, 13, 13, 12, 12, 12, 13, 14, 15, 12, 12, 12, 16]):
        col_letter = get_column_letter(col_idx + 1)
        ws5.column_dimensions[col_letter].width = w

    # -------------------------------------------------------------------------
    # SHEET 6 — Loan Accounts (Unified detail list, NO page-wise splitting)
    # -------------------------------------------------------------------------
    ws6 = wb.create_sheet('All Loan Accounts')
    ws6.sheet_view.showGridLines = False
    ws6.freeze_panes = 'A3'
    
    # Header title
    ws6.merge_cells('A1:R1')
    hdr(ws6['A1'], f'LOAN ACCOUNTS DETAIL  ({len(accounts)} Accounts)', size=13)
    ws6.row_dimensions[1].height = 32
    
    col_headers = [
        '#', 'Lender Name', 'Account Number', 'Account Type', 'Status',
        'Date Opened', 'Last Reported',
        'Current Balance', 'Overdue Amount', 'Credit Limit/Sanctioned', 'High Credit',
        'Asset Classification', 'Security',
        'Interest Rate', 'Installment Amount',
        'Write-off Amount', 'Ownership',
        'DPD / Payment History'
    ]
    for j, h in enumerate(col_headers):
        hdr(ws6.cell(2, j + 1), h, bg=MED_BLUE, size=9)
    ws6.row_dimensions[2].height = 28
    
    def row_bg(status, overdue_amt):
        status_lower = str(status).lower()
        od_val = safe_num(overdue_amt)
        
        if 'write' in status_lower or 'written' in status_lower:
            return RED_BG
        if 'settled' in status_lower:
            return ORANGE_BG
        if od_val > 0:
            return ORANGE_BG
        if 'active' in status_lower:
            return GREEN_BG
        return LIGHT_GRAY
        
    for i, acct in enumerate(accounts):
        bg = row_bg(acct.get("Status"), acct.get("Overdue Amount"))
        rn = i + 3
        ws6.row_dimensions[rn].height = 22
        
        # Values
        curr_bal_val = to_float(acct.get("Current Balance"))
        overdue_val = to_float(acct.get("Overdue Amount"))
        limit_val = to_float(acct.get("Credit Limit/Sanctioned"))
        high_cred_val = to_float(acct.get("High Credit"))
        installment_val = to_float(acct.get("Installment Amount"))
        wo_val = to_float(acct.get("Write-off Amount"))
        
        row_vals = [
            i + 1,
            acct.get("Lender"),
            acct.get("Account Number"),
            acct.get("Account Type"),
            acct.get("Status"),
            acct.get("Date Opened"),
            acct.get("Last Reported"),
            curr_bal_val,
            overdue_val,
            limit_val,
            high_cred_val,
            acct.get("Asset Classification"),
            acct.get("Security"),
            acct.get("Interest Rate"),
            installment_val,
            wo_val,
            acct.get("Ownership"),
            acct.get("DPD History / Payment History")
        ]
        for j, v in enumerate(row_vals):
            c = ws6.cell(rn, j + 1)
            c.value = safe(v)
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
            
            # Format numbers as currency if float (cols: Current Bal, Overdue, Limit, High Credit, Installment, Write-off)
            if j in [7, 8, 9, 10, 14, 15] and isinstance(v, float):
                c.number_format = '₹#,##,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    # Widths for loan accounts
    for col_idx, w in enumerate([4, 20, 15, 18, 12, 13, 13, 15, 15, 15, 13, 16, 12, 12, 14, 14, 14, 24]):
        col_letter = get_column_letter(col_idx + 1)
        ws6.column_dimensions[col_letter].width = w

    # Legend
    legend_row = len(accounts) + 5
    ws6.merge_cells(f'A{legend_row}:R{legend_row}')
    hdr(ws6[f'A{legend_row}'], 'LEGEND', bg=MED_BLUE, size=9)
    ws6.row_dimensions[legend_row].height = 20
    
    legend = [
        (GREEN_BG,    'Active — Good Standing'),
        (RED_BG,      'Written-off / Default'),
        (ORANGE_BG,   'Overdue Balance / Settled'),
        (LIGHT_GRAY,  'Closed'),
    ]
    for li, (bg, desc) in enumerate(legend):
        rr = legend_row + 1 + li
        ws6.row_dimensions[rr].height = 18
        ws6.cell(rr, 1).fill   = PatternFill('solid', start_color=bg)
        ws6.cell(rr, 1).border = thin_border()
        
        ws6.merge_cells(f'B{rr}:E{rr}')
        c           = ws6.cell(rr, 2)
        c.value     = desc
        c.font      = Font(name='Arial', size=9)
        c.alignment = Alignment(vertical='center')
        
        # Apply borders to legend description merged cells
        for col_idx in range(2, 6):
            ws6.cell(rr, col_idx).border = thin_border()

    # -------------------------------------------------------------------------
    # SHEET 7 — Lender Inquiries
    # -------------------------------------------------------------------------
    ws7 = wb.create_sheet('Lender Inquiries')
    ws7.sheet_view.showGridLines = False
    
    ws7.merge_cells('A1:E1')
    hdr(ws7['A1'], f'LENDER INQUIRIES  ({len(inquiries)} Inquiries)', size=13)
    ws7.row_dimensions[1].height = 32
    
    inq_headers = ['#', 'Inquiry Date', 'Inquirer Name', 'Purpose / Loan Type', 'Requested Amount']
    for j, h in enumerate(inq_headers):
        hdr(ws7.cell(2, j + 1), h, bg=MED_BLUE, size=10)
    ws7.row_dimensions[2].height = 24
    
    for idx, inq in enumerate(inquiries):
        rn = idx + 3
        ws7.row_dimensions[rn].height = 18
        
        amt_inq_val = to_float(inq.get("Amount"))
        
        row_vals = [
            idx + 1,
            inq.get("Inquiry Date"),
            inq.get("Inquirer"),
            inq.get("Purpose"),
            amt_inq_val
        ]
        for j, v in enumerate(row_vals):
            c = ws7.cell(rn, j + 1)
            c.value = safe(v)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal='center' if j != 2 and j != 3 else 'left', vertical='center')
            c.border = thin_border()
            
            if j == 4 and isinstance(v, float):
                c.number_format = '₹#,##,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
                
    for col_idx, w in enumerate([4, 15, 25, 25, 18]):
        col_letter = get_column_letter(col_idx + 1)
        ws7.column_dimensions[col_letter].width = w

    # -------------------------------------------------------------------------
    # SHEET 8 — Raw Extracted Data (all fields, for transparency)
    # -------------------------------------------------------------------------
    ws8 = wb.create_sheet('Raw Extracted Data')
    ws8.sheet_view.showGridLines = False

    ws8.merge_cells('A1:H1')
    hdr(ws8['A1'], 'RAW EXTRACTED DATA — ALL PARSED FIELDS', size=13)
    ws8.row_dimensions[1].height = 32

    # Personal info
    ws8.merge_cells('A3:H3')
    hdr(ws8['A3'], 'PERSONAL INFORMATION (RAW)', bg=MED_BLUE)
    pi_fields = list(personal.items())
    for idx, (k, v) in enumerate(pi_fields):
        r = 4 + idx
        ws8.row_dimensions[r].height = 18
        lbl(ws8[f'A{r}'], str(k))
        ws8.merge_cells(f'B{r}:H{r}')
        val(ws8[f'B{r}'], str(v) if v else '-')
    bdr(ws8, 3, 3 + len(pi_fields), 1, 8)

    # All accounts raw dump
    if accounts:
        all_keys = list(accounts[0].keys())
        raw_start = 5 + len(pi_fields) + 2
        ws8.merge_cells(f'A{raw_start}:H{raw_start}')
        hdr(ws8[f'A{raw_start}'], f'ALL ACCOUNT FIELDS ({len(accounts)} accounts)', bg=MED_BLUE)
        ws8.row_dimensions[raw_start].height = 22

        # Header row
        for j, k in enumerate(all_keys):
            hdr(ws8.cell(raw_start + 1, j + 1), k, bg=MED_BLUE, size=8)
        ws8.row_dimensions[raw_start + 1].height = 22

        # Data rows
        for i, acct in enumerate(accounts):
            rn = raw_start + 2 + i
            bg = GREEN_BG if 'active' in str(acct.get('Status', '')).lower() else LIGHT_GRAY
            ws8.row_dimensions[rn].height = 18
            for j, k in enumerate(all_keys):
                c = ws8.cell(rn, j + 1)
                c.value = safe(acct.get(k, '-'))
                c.font = Font(name='Arial', size=8)
                c.fill = PatternFill('solid', start_color=bg)
                c.alignment = Alignment(vertical='center', wrap_text=True)
                c.border = thin_border()

        # Auto-widths for raw sheet
        for col_idx in range(len(all_keys)):
            ws8.column_dimensions[get_column_letter(col_idx + 1)].width = 18

    # -------------------------------------------------------------------------
    # Number every sheet tab serially: '1. ', '2. ', … (pages in serial order)
    # -------------------------------------------------------------------------
    for _i, _ws in enumerate(wb.worksheets, start=1):
        _ws.title = f"{_i}. {_ws.title}"[:31]

    # ── Save ──────────────────────────────────────────────────────────────────
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()
